from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


SHEET_NAMES = (
    "Income statement",
    "Balance sheet",
    "Cash flow statement",
    "Valuation",
)


def _resolve_defined_names(workbook: Workbook) -> dict[str, str]:
    resolved_map: dict[str, list[str]] = {}
    defined_names = workbook.defined_names
    try:
        items = list(defined_names.items())
    except AttributeError:
        items = []
        for defined_name in getattr(defined_names, "definedName", []):
            items.append((defined_name.name, defined_name))

    for name, defined_name in items:
        entries = defined_name if isinstance(defined_name, list) else [defined_name]
        for entry in entries:
            try:
                destinations = list(entry.destinations)
            except AttributeError:
                continue
            if not destinations:
                continue
            parts: list[str] = []
            for sheet_title, coord in destinations:
                parts.append(f"{sheet_title}!{coord}")
            resolved_map.setdefault(name, []).extend(parts)

    return {name: ", ".join(coords) for name, coords in resolved_map.items()}


def _collect_formulas(worksheet: Worksheet) -> list[str]:
    lines: list[str] = []
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.data_type == "f" and cell.value:
                if isinstance(cell.value, ArrayFormula):
                    ref = cell.value.ref or cell.coordinate
                    lines.append(f"{cell.coordinate} ({ref}): {cell.value.text}")
                else:
                    lines.append(f"{cell.coordinate}: ={cell.value}")
    return lines


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    workbook_path = root / "docs" / "industrial-model-20250616.xlsm"
    output_path = root / "docs" / "forecast_logic_reference.txt"

    workbook = load_workbook(workbook_path, data_only=False, keep_vba=True)
    defined_names = _resolve_defined_names(workbook)

    lines: list[str] = []
    for sheet_name in SHEET_NAMES:
        if sheet_name not in workbook.sheetnames:
            lines.append(f"[Missing sheet] {sheet_name}")
            continue
        worksheet = workbook[sheet_name]
        lines.append(f"[Sheet] {sheet_name}")
        lines.extend(_collect_formulas(worksheet))
        lines.append("")

    lines.append("[Defined Names]")
    for name in sorted(defined_names):
        lines.append(f"{name}: {defined_names[name]}")
    lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
